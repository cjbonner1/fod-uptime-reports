import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict

wb = Workbook()

FN = "Aptos Narrow"
title_font = Font(name=FN, size=24, bold=True)
section_font = Font(name=FN, size=11, bold=True)
body_font = Font(name=FN, size=11)
note_font = Font(name=FN, size=10, italic=True, color="666666")
hdr_fill = PatternFill("solid", fgColor="2F5496")
hdr_font_w = Font(name=FN, bold=True, color="FFFFFF", size=11)
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
light_fill = PatternFill("solid", fgColor="D6E4F0")
green_font = Font(name=FN, size=11, bold=True, color="006100")

# ============================================================
# 2026 DATA (UPDATED: includes full Mar 2026 + Apr partial)
# ============================================================
raw_2026 = [
    # Apr 2026 (partial - through Apr 16)
    ("2026-04-10 20:56","2026-04-10 22:32","Maintenance","SAST Aviator / ams-sast-aviator"),
    ("2026-04-09 21:04","2026-04-09 21:34","Maintenance","SAST Aviator / eu-sast-aviator"),
    ("2026-04-09 19:11","2026-04-09 19:13","Outage","SAST Aviator / eu-sast-aviator"),
    ("2026-04-09 18:04","2026-04-09 21:35","Maintenance","SAST Aviator / eu-sast-aviator"),
    ("2026-04-08 04:10","2026-04-08 04:12","Outage","Fortify on Demand API - FedRAMP"),
    ("2026-04-08 04:09","2026-04-08 04:12","Outage","Fortify on Demand Tenant Portal - FedRAMP"),
    ("2026-04-07 18:03","2026-04-07 19:21","Service Degradation","Fortify on Demand Tenant Portal - AMS"),
    ("2026-04-02 00:27","2026-04-02 01:17","Maintenance","Fortify on Demand Tenant Portal - FedRAMP"),
    ("2026-04-02 00:27","2026-04-02 01:17","Maintenance","Fortify on Demand API - FedRAMP"),
    # Mar 2026 (COMPLETE)
    ("2026-03-26 03:52","2026-03-26 07:15","Maintenance","Fortify on Demand Tenant Portal - AMS"),
    ("2026-03-26 03:52","2026-03-26 07:15","Maintenance","Fortify on Demand API - AMS"),
    ("2026-03-25 23:07","2026-03-26 01:37","Maintenance","Fortify on Demand Tenant Portal - EMEA"),
    ("2026-03-25 23:07","2026-03-26 01:37","Maintenance","Fortify on Demand API - EMEA"),
    ("2026-03-25 14:07","2026-03-25 16:10","Maintenance","Fortify on Demand Tenant Portal - APAC"),
    ("2026-03-25 14:07","2026-03-25 16:10","Maintenance","Fortify on Demand API - APAC"),
    ("2026-03-25 14:07","2026-03-25 16:10","Maintenance","Fortify on Demand Tenant Portal - SGP"),
    ("2026-03-25 14:07","2026-03-25 16:10","Maintenance","Fortify on Demand API - SGP"),
    ("2026-03-20 21:04","2026-03-20 22:34","Maintenance","SAST Aviator / ams-sast-aviator"),
    ("2026-03-20 17:44","2026-03-20 18:26","Maintenance","SAST Aviator / eu-sast-aviator"),
    ("2026-03-18 19:11","2026-03-18 20:50","Maintenance","Fortify on Demand Tenant Portal - EMEA"),
    ("2026-03-18 19:11","2026-03-18 20:50","Maintenance","Fortify on Demand API - EMEA"),
    ("2026-03-18 00:43","2026-03-18 00:46","Outage","SAST Aviator / ams-sast-aviator"),
    ("2026-03-17 22:16","2026-03-18 01:34","Maintenance","Fortify on Demand Tenant Portal - AMS"),
    ("2026-03-17 22:16","2026-03-18 01:34","Maintenance","Fortify on Demand API - AMS"),
    ("2026-03-17 19:46","2026-03-17 19:48","Outage","SAST Aviator / eu-sast-aviator"),
    ("2026-03-13 00:37","2026-03-13 00:41","Outage","SAST Aviator / ams-sast-aviator"),
    ("2026-03-11 23:50","2026-03-11 23:54","Outage","SAST Aviator / ams-sast-aviator"),
    ("2026-03-05 05:01","2026-03-05 07:45","Maintenance","Fortify on Demand Tenant Portal - AMS"),
    ("2026-03-05 05:01","2026-03-05 07:45","Maintenance","Fortify on Demand API - AMS"),
    ("2026-03-05 01:20","2026-03-05 02:45","Maintenance","Fortify on Demand Tenant Portal - FedRAMP"),
    ("2026-03-05 01:20","2026-03-05 02:45","Maintenance","Fortify on Demand API - FedRAMP"),
    ("2026-03-04 23:49","2026-03-05 01:16","Maintenance","Fortify on Demand Tenant Portal - EMEA"),
    ("2026-03-04 23:49","2026-03-05 01:16","Maintenance","Fortify on Demand API - EMEA"),
    ("2026-03-04 14:35","2026-03-04 16:10","Maintenance","Fortify on Demand Tenant Portal - APAC"),
    ("2026-03-04 14:35","2026-03-04 16:10","Maintenance","Fortify on Demand API - APAC"),
    ("2026-03-04 14:35","2026-03-04 16:10","Maintenance","Fortify on Demand Tenant Portal - SGP"),
    ("2026-03-04 14:35","2026-03-04 16:10","Maintenance","Fortify on Demand API - SGP"),
    ("2026-03-02 16:51","2026-03-02 17:02","Outage","Fortify on Demand Tenant Portal - FedRAMP"),
    ("2026-03-02 16:51","2026-03-02 16:56","Outage","Fortify on Demand API - FedRAMP"),
    ("2026-03-02 14:44","2026-03-02 15:15","Outage","Fortify on Demand Tenant Portal - FedRAMP"),
    ("2026-03-02 14:42","2026-03-02 14:44","Outage","Fortify on Demand API - FedRAMP"),
    ("2026-03-02 14:41","2026-03-02 14:42","Outage","Fortify on Demand Tenant Portal - FedRAMP"),
    # Feb 2026
    ("2026-02-26 16:39","2026-02-26 16:44","Outage","Fortify on Demand API - SGP"),
    ("2026-02-20 15:28","2026-02-20 18:55","Maintenance","Fortify on Demand Tenant Portal - APAC"),
    ("2026-02-20 15:28","2026-02-20 18:55","Maintenance","Fortify on Demand API - APAC"),
    ("2026-02-18 19:36","2026-02-18 22:15","Maintenance","Fortify on Demand Tenant Portal - AMS"),
    ("2026-02-18 19:36","2026-02-18 22:15","Maintenance","Fortify on Demand API - AMS"),
    ("2026-02-15 19:46","2026-02-15 19:48","Outage","SAST Aviator / eu-sast-aviator"),
    ("2026-02-14 22:07","2026-02-14 22:09","Outage","SAST Aviator / ams-sast-aviator"),
    ("2026-02-05 16:28","2026-02-05 21:10","Maintenance","Fortify on Demand Tenant Portal - SGP"),
    ("2026-02-05 16:28","2026-02-05 21:10","Maintenance","Fortify on Demand API - SGP"),
    ("2026-02-05 04:56","2026-02-05 07:03","Maintenance","Fortify on Demand Tenant Portal - AMS"),
    ("2026-02-05 04:56","2026-02-05 07:03","Maintenance","Fortify on Demand API - AMS"),
    ("2026-02-05 01:57","2026-02-05 03:35","Maintenance","Fortify on Demand Tenant Portal - FedRAMP"),
    ("2026-02-05 01:57","2026-02-05 03:35","Maintenance","Fortify on Demand API - FedRAMP"),
    ("2026-02-04 23:58","2026-02-05 01:26","Maintenance","Fortify on Demand Tenant Portal - EMEA"),
    ("2026-02-04 23:58","2026-02-05 01:26","Maintenance","Fortify on Demand API - EMEA"),
    # Jan-Feb bridge
    ("2026-01-30 19:08","2026-02-02 17:42","Maintenance","Fortify on Demand Tenant Portal - EMEA"),
    ("2026-01-30 19:08","2026-02-02 17:42","Maintenance","Fortify on Demand API - EMEA"),
    # Jan 2026
    ("2026-01-27 14:46","2026-01-27 14:47","Outage","Fortify on Demand API - FedRAMP"),
    ("2026-01-27 14:45","2026-01-27 14:46","Outage","Fortify on Demand Tenant Portal - FedRAMP"),
    ("2026-01-23 23:57","2026-01-25 02:32","Maintenance","Fortify on Demand Tenant Portal - AMS"),
    ("2026-01-23 23:57","2026-01-25 02:32","Maintenance","Fortify on Demand API - AMS"),
    ("2026-01-16 16:24","2026-01-20 03:26","Maintenance","SAST Aviator / eu-sast-aviator"),
    ("2026-01-16 15:14","2026-01-16 15:15","Outage","SAST Aviator / eu-sast-aviator"),
    ("2026-01-15 21:13","2026-01-16 15:03","Maintenance","SAST Aviator / ams-sast-aviator"),
    ("2026-01-15 04:24","2026-01-15 04:25","Outage","Fortify on Demand - Vulncat"),
    ("2026-01-10 15:08","2026-01-10 15:10","Outage","SAST Aviator / eu-sast-aviator"),
    ("2026-01-08 04:54","2026-01-08 07:14","Maintenance","Fortify on Demand Tenant Portal - AMS"),
    ("2026-01-08 04:54","2026-01-08 07:14","Maintenance","Fortify on Demand API - AMS"),
    ("2026-01-08 01:13","2026-01-08 03:30","Maintenance","Fortify on Demand Tenant Portal - FedRAMP"),
    ("2026-01-08 01:13","2026-01-08 03:30","Maintenance","Fortify on Demand API - FedRAMP"),
    ("2026-01-08 00:24","2026-01-08 01:02","Maintenance","Fortify on Demand Tenant Portal - EMEA"),
    ("2026-01-08 00:24","2026-01-08 01:02","Maintenance","Fortify on Demand API - EMEA"),
]

parsed = []
for s, e, typ, comp in raw_2026:
    sd = datetime.strptime(s, "%Y-%m-%d %H:%M")
    ed = datetime.strptime(e, "%Y-%m-%d %H:%M")
    dur = max((ed - sd).total_seconds() / 60, 0)
    parsed.append((sd, ed, dur, typ, comp))

outages_only = [(sd,ed,dur,typ,comp) for sd,ed,dur,typ,comp in parsed if typ == "Outage"]

# De-dup: group simultaneous outages (start +/- 1 min)
outages_sorted = sorted(outages_only, key=lambda x: x[0])
grouped = []
i = 0
while i < len(outages_sorted):
    sd, ed, dur, typ, comp = outages_sorted[i]
    g_comps = [comp]; g_dur = dur; g_ed = ed
    j = i + 1
    while j < len(outages_sorted):
        sd2, ed2, dur2, typ2, comp2 = outages_sorted[j]
        if abs((sd2 - sd).total_seconds()) <= 60:
            g_comps.append(comp2); g_dur = max(g_dur, dur2); g_ed = max(g_ed, ed2); j += 1
        else:
            break
    grouped.append({'start': sd, 'end': g_ed, 'duration': g_dur, 'components': g_comps})
    i = j

# Monthly/daily stats
daily = defaultdict(lambda: {'count': 0, 'minutes': 0})
monthly = defaultdict(lambda: {'count': 0, 'minutes': 0})
for inc in grouped:
    d = inc['start'].date()
    m = inc['start'].strftime("%Y-%m")
    daily[d]['count'] += 1
    daily[d]['minutes'] += inc['duration']
    monthly[m]['count'] += 1
    monthly[m]['minutes'] += inc['duration']

# Per-component outage minutes (no cross-component de-dup, for Regional Breakout tab)
# Bucket mapping: "contains" string -> bucket name
bucket_map = {
    "Tenant Portal - FedRAMP": "FedRAMP Portal",
    "API - FedRAMP": "FedRAMP API",
    "Tenant Portal - AMS": "AMS Portal",
    "API - AMS": "AMS API",
    "Tenant Portal - EMEA": "EMEA Portal",
    "API - EMEA": "EMEA API",
    "Tenant Portal - APAC": "APAC Portal",
    "API - APAC": "APAC API",
    "Tenant Portal - SGP": "SGP Portal",
    "API - SGP": "SGP API",
    "Tenant Portal - EU": "EU Portal",
    "API - EU": "EU API",
    "eu-sast-aviator": "EU SAST Aviator",
    "ams-sast-aviator": "AMS SAST Aviator",
    "Debricked Fortify Integration": "Debricked Fortify Integration",
    "Vulncat": "Vulncat",
}

def resolve_bucket(component_name):
    for key, bucket in bucket_map.items():
        if key in component_name:
            return bucket
    return component_name

comp_monthly_outage = defaultdict(lambda: defaultdict(float))
for sd, ed, dur, typ, comp in outages_only:
    m = sd.strftime("%Y-%m")
    bkt = resolve_bucket(comp)
    comp_monthly_outage[bkt][m] += dur

days_map = {1:31, 2:28, 3:31, 4:30, 5:31, 6:30, 7:31, 8:31, 9:30, 10:31, 11:30, 12:31}

def month_minutes(m_str):
    mo = int(m_str[5:])
    return days_map[mo] * 1440

# ============================================================
# TAB 1: 2026 Executive Summary
# ============================================================
ws = wb.active
ws.title = "2026 Executive Summary"
ws.column_dimensions['A'].width = 42
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 85

row = 1
ws.cell(row=row, column=1, value="Fortify Service Uptime Report").font = title_font
row += 2
ws.cell(row=row, column=1, value="2026 Year-to-Date").font = section_font
row += 2

# Exec summary
ws.cell(row=row, column=1, value="Executive Summary").font = section_font
ws.cell(row=row, column=7, value="Key Observations").font = section_font
row += 2

# YTD stats (Jan + Feb + Mar all complete)
months_complete = ["2026-01", "2026-02", "2026-03"]
ytd_minutes = sum(month_minutes(m) for m in months_complete)
ytd_outage_min = sum(monthly.get(m, {'minutes':0})['minutes'] for m in months_complete)
ytd_outage_count = sum(monthly.get(m, {'count':0})['count'] for m in months_complete)
ytd_uptime = 1 - (ytd_outage_min / ytd_minutes)

ws.cell(row=row, column=1, value="Report Period: Calendar Year 2026").font = body_font
ws.cell(row=row, column=7, value=f"1. January: {monthly['2026-01']['count']} outages ({monthly['2026-01']['minutes']:.0f} min); February: {monthly['2026-02']['count']} outages ({monthly['2026-02']['minutes']:.0f} min); March: {monthly.get('2026-03',{'count':0})['count']} outages ({monthly.get('2026-03',{'minutes':0})['minutes']:.0f} min)").font = body_font
row += 1
ws.cell(row=row, column=1, value=f"Completed Months: January - March 2026 (April in progress)").font = body_font
ws.cell(row=row, column=7, value=f"2. March 2 saw FedRAMP Portal/API hit with multiple back-to-back outages totaling 44 min, the highest single-day downtime in 2026").font = body_font
row += 1
ws.cell(row=row, column=1, value=f"YTD Outages: {ytd_outage_count} incidents (completed months)").font = body_font
ws.cell(row=row, column=7, value="3. SAST Aviator (EU/AMS) logged outages on Mar 11, 13, 17, and 18, consistent with recurring brief disruption pattern from prior months").font = body_font
row += 1
ws.cell(row=row, column=1, value=f"YTD Outage Time: {ytd_outage_min:.0f} minutes ({ytd_outage_min/60:.2f} hours) (completed months)").font = body_font
ws.cell(row=row, column=7, value=f"4. YTD uptime of {ytd_uptime*100:.4f}% continues to exceed the 99.9% SLA target across all three completed months").font = body_font
row += 1
ws.cell(row=row, column=1, value=f"YTD Uptime: {ytd_uptime*100:.4f}% (completed months only)").font = body_font
ws.cell(row=row, column=7, value="5. SAST Aviator (EU/AMS) accounts for the majority of incidents across all months; FedRAMP had the single worst day on Mar 2").font = body_font
row += 2

# Monthly Summary Table
ws.cell(row=row, column=1, value="Monthly Uptime Summary").font = section_font
row += 1
for col, h in enumerate(["Month", "Outage Count", "Outage Minutes", "Outage Hours", "Uptime Percentage"], 1):
    c = ws.cell(row=row, column=col, value=h)
    c.font = hdr_font_w; c.fill = hdr_fill; c.border = thin_border; c.alignment = Alignment(horizontal="center")
row += 1

for m in ["2026-01","2026-02","2026-03","2026-04","2026-05","2026-06","2026-07","2026-08","2026-09","2026-10","2026-11","2026-12"]:
    stats = monthly.get(m, {'count':0, 'minutes':0})
    tm = month_minutes(m)
    is_future = m > "2026-04"
    is_partial = m == "2026-04"

    ws.cell(row=row, column=1, value=m).font = body_font
    ws.cell(row=row, column=1).border = thin_border

    if is_future:
        for col in range(2,6):
            ws.cell(row=row, column=col).font = note_font
            ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=2, value="-").font = note_font
        ws.cell(row=row, column=3, value="-").font = note_font
        ws.cell(row=row, column=4, value="-").font = note_font
        ws.cell(row=row, column=5, value="-").font = note_font
    elif is_partial:
        partial_count = stats['count'] if stats['count'] > 0 else 0
        partial_min = int(stats['minutes'])
        partial_up = 1 - (stats['minutes']/tm) if stats['minutes'] > 0 else 1
        ws.cell(row=row, column=2, value=f"{partial_count}*").font = note_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3, value=f"{partial_min}*").font = note_font
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=4, value=f"{round(stats['minutes']/60,2)}*").font = note_font
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=5, value=partial_up).font = note_font
        ws.cell(row=row, column=5).number_format = "0.0000%"
        ws.cell(row=row, column=5).border = thin_border
    else:
        ws.cell(row=row, column=2, value=stats['count']).font = body_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3, value=int(stats['minutes'])).font = body_font
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=4, value=round(stats['minutes']/60, 2)).font = body_font
        ws.cell(row=row, column=4).border = thin_border
        up = 1 - (stats['minutes']/tm) if stats['minutes'] > 0 else 1
        ws.cell(row=row, column=5, value=up).font = body_font
        ws.cell(row=row, column=5).number_format = "0.0000%"
        ws.cell(row=row, column=5).border = thin_border
    row += 1

# YTD Totals
ws.cell(row=row, column=1, value="YTD TOTAL").font = Font(name=FN, bold=True, size=11)
ws.cell(row=row, column=1).fill = light_fill; ws.cell(row=row, column=1).border = thin_border
ws.cell(row=row, column=2, value=ytd_outage_count).font = Font(name=FN, bold=True, size=11)
ws.cell(row=row, column=2).fill = light_fill; ws.cell(row=row, column=2).border = thin_border
ws.cell(row=row, column=3, value=int(ytd_outage_min)).font = Font(name=FN, bold=True, size=11)
ws.cell(row=row, column=3).fill = light_fill; ws.cell(row=row, column=3).border = thin_border
ws.cell(row=row, column=4, value=round(ytd_outage_min/60, 2)).font = Font(name=FN, bold=True, size=11)
ws.cell(row=row, column=4).fill = light_fill; ws.cell(row=row, column=4).border = thin_border
ws.cell(row=row, column=5, value=ytd_uptime).font = Font(name=FN, bold=True, size=11, color="006100")
ws.cell(row=row, column=5).number_format = "0.0000%"
ws.cell(row=row, column=5).fill = light_fill; ws.cell(row=row, column=5).border = thin_border
row += 1
ws.cell(row=row, column=1, value="* April 2026 is in progress. Partial data through April 16, 2026. Uptime % will change as month completes.").font = note_font
row += 2

# SLA block
ws.cell(row=row, column=1, value="2026 SLA Performance (Completed Months: Jan - Mar)").font = section_font
row += 2
avg_dur = ytd_outage_min / max(ytd_outage_count, 1)
sla = [
    f"Total Period Minutes (completed): {ytd_minutes:,}",
    f"Total Outage Minutes: {ytd_outage_min:.0f}",
    f"Uptime Minutes: {ytd_minutes - ytd_outage_min:,.0f}",
    f"Uptime Percentage: {ytd_uptime*100:.4f}%",
    f"Outage Count: {ytd_outage_count} incidents",
    f"Average Outage Duration: {avg_dur:.2f} minutes",
    f"SLA Target: 99.9%",
    f"SLA Status: MEETING TARGET" if ytd_uptime >= 0.999 else f"SLA Status: BELOW TARGET",
]
for line in sla:
    ws.cell(row=row, column=1, value=line).font = body_font
    row += 1


# ============================================================
# TAB 2: Regional Breakout (per-component uptime)
# ============================================================
ws_rb = wb.create_sheet("Regional Breakout")
ws_rb.column_dimensions['A'].width = 30

all_months = ["2026-01","2026-02","2026-03","2026-04","2026-05","2026-06",
              "2026-07","2026-08","2026-09","2026-10","2026-11","2026-12"]
month_labels = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
# Completed months for YTD (same as exec summary)
ytd_months = months_complete

rb_row = 1
ws_rb.cell(row=rb_row, column=1, value="Regional Breakout: Per-Component Uptime").font = title_font
rb_row += 2

def write_rb_section(start_row, section_title, components):
    """Write a section of the Regional Breakout tab. Returns next available row."""
    r = start_row
    ws_rb.cell(row=r, column=1, value=section_title).font = section_font
    r += 1

    # Header row: Component | Jan | Feb | ... | Dec | YTD
    headers = ["Component"] + month_labels + ["YTD"]
    for ci, h in enumerate(headers):
        c = ws_rb.cell(row=r, column=ci + 1, value=h)
        c.font = hdr_font_w
        c.fill = hdr_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
    # Set month column widths
    for ci in range(2, 15):
        ws_rb.column_dimensions[get_column_letter(ci)].width = 10
    r += 1

    for comp_name in components:
        ws_rb.cell(row=r, column=1, value=comp_name).font = body_font
        ws_rb.cell(row=r, column=1).border = thin_border

        ytd_outage = 0.0
        ytd_total = 0.0

        for mi, m_str in enumerate(all_months):
            col = mi + 2
            is_future = m_str > "2026-04"
            is_partial = m_str == "2026-04"
            tm = month_minutes(m_str)
            outage_min = comp_monthly_outage[comp_name].get(m_str, 0.0)

            if is_future:
                ws_rb.cell(row=r, column=col, value="-").font = note_font
                ws_rb.cell(row=r, column=col).border = thin_border
                ws_rb.cell(row=r, column=col).alignment = Alignment(horizontal="center")
            else:
                up = 1 - (outage_min / tm) if tm > 0 else 1
                cell = ws_rb.cell(row=r, column=col, value=up)
                cell.number_format = "0.0000%"
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if is_partial:
                    cell.font = note_font
                elif up >= 0.999:
                    cell.font = green_font
                else:
                    cell.font = body_font

            # Accumulate YTD from completed months only
            if m_str in ytd_months:
                ytd_outage += outage_min
                ytd_total += tm

        # YTD column (col 14)
        if ytd_total > 0:
            ytd_up = 1 - (ytd_outage / ytd_total)
            ytd_cell = ws_rb.cell(row=r, column=14, value=ytd_up)
            ytd_cell.number_format = "0.0000%"
            ytd_cell.border = thin_border
            ytd_cell.alignment = Alignment(horizontal="center")
            if ytd_up >= 0.999:
                ytd_cell.font = green_font
            else:
                ytd_cell.font = body_font
        else:
            ws_rb.cell(row=r, column=14, value="-").font = note_font
            ws_rb.cell(row=r, column=14).border = thin_border

        r += 1
    return r + 1  # blank row after section

core_services = [
    "FedRAMP Portal", "FedRAMP API",
    "AMS Portal", "AMS API",
    "EMEA Portal", "EMEA API",
    "APAC Portal", "APAC API",
    "SGP Portal", "SGP API",
    "EU Portal", "EU API",
]
sast_aviator = ["AMS SAST Aviator", "EU SAST Aviator"]
other_services = ["Debricked Fortify Integration", "Vulncat"]

rb_row = write_rb_section(rb_row, "Core Services", core_services)
rb_row = write_rb_section(rb_row, "SAST Aviator", sast_aviator)
rb_row = write_rb_section(rb_row, "Other Services", other_services)

# Footer note
ws_rb.cell(row=rb_row, column=1, value="* April 2026 is in progress. Partial data shown in italics.").font = note_font
rb_row += 1
ws_rb.cell(row=rb_row, column=1, value="Uptime = (total_month_minutes - outage_minutes) / total_month_minutes. Only Outage events counted. No cross-component de-duplication.").font = note_font

# ============================================================
# TAB 3: Jan 2026 (detail tab - unchanged)
# ============================================================
ws_jan = wb.create_sheet("Jan 2026")
ws_jan.column_dimensions['A'].width = 42
ws_jan.column_dimensions['B'].width = 16
ws_jan.column_dimensions['C'].width = 18
ws_jan.column_dimensions['D'].width = 18
ws_jan.column_dimensions['E'].width = 18
ws_jan.column_dimensions['G'].width = 85

r = 1
ws_jan.cell(row=r, column=1, value="January 2026 - Monthly Detail").font = title_font
r += 2

jan_min = 44640
jan_stats = monthly.get("2026-01", {'count':0,'minutes':0})
jan_up = 1 - (jan_stats['minutes'] / jan_min)

ws_jan.cell(row=r, column=1, value="Monthly Summary").font = section_font
r += 2
for label, val in [("Period:", "January 1 - 31, 2026"), ("Total Minutes:", f"{jan_min:,}"), ("Outage Count:", str(jan_stats['count'])), ("Outage Minutes:", str(int(jan_stats['minutes']))), ("Uptime %:", f"{jan_up*100:.4f}%"), ("SLA Status:", "MEETING TARGET (99.9%)")]:
    ws_jan.cell(row=r, column=1, value=label).font = section_font
    ws_jan.cell(row=r, column=2, value=val).font = body_font
    r += 1
r += 1

ws_jan.cell(row=r, column=1, value="Daily Outage Report").font = section_font
r += 1
for col, h in enumerate(["Date", "Outage Count", "Outage Minutes", "Uptime Percentage"], 1):
    c = ws_jan.cell(row=r, column=col, value=h)
    c.font = hdr_font_w; c.fill = hdr_fill; c.border = thin_border; c.alignment = Alignment(horizontal="center")
r += 1

jan_daily = {d: s for d, s in daily.items() if d.month == 1 and d.year == 2026}
for d, stats in sorted(jan_daily.items(), key=lambda x: -x[1]['minutes']):
    ws_jan.cell(row=r, column=1, value=datetime(d.year, d.month, d.day)).font = body_font
    ws_jan.cell(row=r, column=1).number_format = "mm-dd-yy"
    ws_jan.cell(row=r, column=1).border = thin_border
    ws_jan.cell(row=r, column=2, value=stats['count']).font = body_font
    ws_jan.cell(row=r, column=2).border = thin_border
    ws_jan.cell(row=r, column=3, value=int(stats['minutes'])).font = body_font
    ws_jan.cell(row=r, column=3).border = thin_border
    ws_jan.cell(row=r, column=4, value=1 - (stats['minutes']/1440)).font = body_font
    ws_jan.cell(row=r, column=4).number_format = "0.00%"
    ws_jan.cell(row=r, column=4).border = thin_border
    r += 1
r += 1

ws_jan.cell(row=r, column=1, value="Services Affected").font = section_font
r += 2
jan_svcs = set()
for inc in grouped:
    if inc['start'].month == 1 and inc['start'].year == 2026:
        for c in inc['components']: jan_svcs.add(c)
for svc in sorted(jan_svcs):
    ws_jan.cell(row=r, column=1, value=svc).font = body_font
    r += 1
r += 2

ws_jan.cell(row=r, column=1, value="Detailed Incidents").font = section_font
r += 2
jan_incs = sorted([inc for inc in grouped if inc['start'].month == 1 and inc['start'].year == 2026], key=lambda x: -x['duration'])
for idx, inc in enumerate(jan_incs, 1):
    comps = inc['components']
    dur = inc['duration']
    date_fmt = inc['start'].strftime("%B %d, %Y")
    start_fmt = inc['start'].strftime("%-I:%M %p GMT")
    end_fmt = inc['end'].strftime("%-I:%M %p GMT")
    if len(comps) > 1:
        title = f"{idx}. {date_fmt} - Multiple Outages ({dur:.0f} minutes total)"
    else:
        title = f"{idx}. {date_fmt} - Outage ({dur:.0f} minute{'s' if dur != 1 else ''})"
    ws_jan.cell(row=r, column=1, value=title).font = section_font
    r += 2
    if len(comps) > 1:
        for ci, c in enumerate(comps, 1):
            ws_jan.cell(row=r, column=1, value=f"Service {ci}: {c}").font = body_font
            r += 1
    else:
        ws_jan.cell(row=r, column=1, value=f"Service: {comps[0]}").font = body_font
        r += 1
    ws_jan.cell(row=r, column=1, value=f"Time: {start_fmt}-{end_fmt}").font = body_font
    r += 1
    ws_jan.cell(row=r, column=1, value="Impact: Brief service disruption").font = body_font
    r += 2

# ============================================================
# TAB 4: Feb 2026 (FINALIZED - complete month)
# ============================================================
ws_feb = wb.create_sheet("Feb 2026")
ws_feb.column_dimensions['A'].width = 42
ws_feb.column_dimensions['B'].width = 16
ws_feb.column_dimensions['C'].width = 18
ws_feb.column_dimensions['D'].width = 18
ws_feb.column_dimensions['E'].width = 18
ws_feb.column_dimensions['G'].width = 85

r = 1
ws_feb.cell(row=r, column=1, value="February 2026 - Monthly Detail").font = title_font
r += 2

feb_min = 40320
feb_stats = monthly.get("2026-02", {'count':0,'minutes':0})
feb_up = 1 - (feb_stats['minutes'] / feb_min)

ws_feb.cell(row=r, column=1, value="Monthly Summary").font = section_font
r += 2
for label, val in [("Period:", "February 1 - 28, 2026"), ("Total Minutes:", f"{feb_min:,}"), ("Outage Count:", str(feb_stats['count'])), ("Outage Minutes:", str(int(feb_stats['minutes']))), ("Uptime %:", f"{feb_up*100:.4f}%"), ("SLA Status:", "MEETING TARGET (99.9%)")]:
    ws_feb.cell(row=r, column=1, value=label).font = section_font
    ws_feb.cell(row=r, column=2, value=val).font = body_font
    r += 1
r += 1

# Daily table
ws_feb.cell(row=r, column=1, value="Daily Outage Report").font = section_font
r += 1
for col, h in enumerate(["Date", "Outage Count", "Outage Minutes", "Uptime Percentage"], 1):
    c = ws_feb.cell(row=r, column=col, value=h)
    c.font = hdr_font_w; c.fill = hdr_fill; c.border = thin_border; c.alignment = Alignment(horizontal="center")
r += 1

feb_daily = {d: s for d, s in daily.items() if d.month == 2 and d.year == 2026}
for d, stats in sorted(feb_daily.items(), key=lambda x: -x[1]['minutes']):
    ws_feb.cell(row=r, column=1, value=datetime(d.year, d.month, d.day)).font = body_font
    ws_feb.cell(row=r, column=1).number_format = "mm-dd-yy"
    ws_feb.cell(row=r, column=1).border = thin_border
    ws_feb.cell(row=r, column=2, value=stats['count']).font = body_font
    ws_feb.cell(row=r, column=2).border = thin_border
    ws_feb.cell(row=r, column=3, value=int(stats['minutes'])).font = body_font
    ws_feb.cell(row=r, column=3).border = thin_border
    ws_feb.cell(row=r, column=4, value=1 - (stats['minutes']/1440)).font = body_font
    ws_feb.cell(row=r, column=4).number_format = "0.00%"
    ws_feb.cell(row=r, column=4).border = thin_border
    r += 1
r += 1

# Service distribution
ws_feb.cell(row=r, column=1, value="Services Affected").font = section_font
r += 2
feb_svcs = set()
for inc in grouped:
    if inc['start'].month == 2 and inc['start'].year == 2026:
        for c in inc['components']: feb_svcs.add(c)
for svc in sorted(feb_svcs):
    ws_feb.cell(row=r, column=1, value=svc).font = body_font
    r += 1
r += 2

# Detailed incidents
ws_feb.cell(row=r, column=1, value="Detailed Incidents").font = section_font
r += 2
feb_incs = sorted([inc for inc in grouped if inc['start'].month == 2 and inc['start'].year == 2026], key=lambda x: -x['duration'])
for idx, inc in enumerate(feb_incs, 1):
    comps = inc['components']
    dur = inc['duration']
    date_fmt = inc['start'].strftime("%B %d, %Y")
    start_fmt = inc['start'].strftime("%-I:%M %p GMT")
    end_fmt = inc['end'].strftime("%-I:%M %p GMT")
    if len(comps) > 1:
        title = f"{idx}. {date_fmt} - Multiple Outages ({dur:.0f} minutes total)"
    else:
        title = f"{idx}. {date_fmt} - Outage ({dur:.0f} minute{'s' if dur != 1 else ''})"
    ws_feb.cell(row=r, column=1, value=title).font = section_font
    r += 2
    if len(comps) > 1:
        for ci, c in enumerate(comps, 1):
            ws_feb.cell(row=r, column=1, value=f"Service {ci}: {c}").font = body_font
            r += 1
    else:
        ws_feb.cell(row=r, column=1, value=f"Service: {comps[0]}").font = body_font
        r += 1
    ws_feb.cell(row=r, column=1, value=f"Time: {start_fmt}-{end_fmt}").font = body_font
    r += 1
    ws_feb.cell(row=r, column=1, value="Impact: Brief service disruption").font = body_font
    r += 2

# ============================================================
# TAB 5: Mar 2026 (FINALIZED - complete month)
# ============================================================
ws_mar = wb.create_sheet("Mar 2026")
ws_mar.column_dimensions['A'].width = 42
ws_mar.column_dimensions['B'].width = 16
ws_mar.column_dimensions['C'].width = 18
ws_mar.column_dimensions['D'].width = 18
ws_mar.column_dimensions['E'].width = 18
ws_mar.column_dimensions['G'].width = 85

r = 1
ws_mar.cell(row=r, column=1, value="March 2026 - Monthly Detail").font = title_font
r += 2

mar_min = 44640
mar_stats = monthly.get("2026-03", {'count':0,'minutes':0})
mar_up = 1 - (mar_stats['minutes'] / mar_min) if mar_stats['minutes'] > 0 else 1

ws_mar.cell(row=r, column=1, value="Monthly Summary").font = section_font
r += 2
for label, val in [
    ("Period:", "March 1 - 31, 2026"),
    ("Total Minutes:", f"{mar_min:,}"),
    ("Outage Count:", str(mar_stats['count'])),
    ("Outage Minutes:", str(int(mar_stats['minutes']))),
    ("Uptime %:", f"{mar_up*100:.4f}%"),
    ("SLA Status:", "MEETING TARGET (99.9%)" if mar_up >= 0.999 else "BELOW TARGET"),
]:
    ws_mar.cell(row=r, column=1, value=label).font = section_font
    ws_mar.cell(row=r, column=2, value=val).font = body_font
    r += 1
r += 1

ws_mar.cell(row=r, column=1, value="Daily Outage Report").font = section_font
r += 1
for col, h in enumerate(["Date", "Outage Count", "Outage Minutes", "Uptime Percentage"], 1):
    c = ws_mar.cell(row=r, column=col, value=h)
    c.font = hdr_font_w; c.fill = hdr_fill; c.border = thin_border; c.alignment = Alignment(horizontal="center")
r += 1

mar_daily = {d: s for d, s in daily.items() if d.month == 3 and d.year == 2026}
for d, stats in sorted(mar_daily.items(), key=lambda x: -x[1]['minutes']):
    ws_mar.cell(row=r, column=1, value=datetime(d.year, d.month, d.day)).font = body_font
    ws_mar.cell(row=r, column=1).number_format = "mm-dd-yy"
    ws_mar.cell(row=r, column=1).border = thin_border
    ws_mar.cell(row=r, column=2, value=stats['count']).font = body_font
    ws_mar.cell(row=r, column=2).border = thin_border
    ws_mar.cell(row=r, column=3, value=int(stats['minutes'])).font = body_font
    ws_mar.cell(row=r, column=3).border = thin_border
    ws_mar.cell(row=r, column=4, value=1 - (stats['minutes']/1440)).font = body_font
    ws_mar.cell(row=r, column=4).number_format = "0.00%"
    ws_mar.cell(row=r, column=4).border = thin_border
    r += 1
r += 1

ws_mar.cell(row=r, column=1, value="Services Affected").font = section_font
r += 2
mar_svcs = set()
for inc in grouped:
    if inc['start'].month == 3 and inc['start'].year == 2026:
        for c in inc['components']: mar_svcs.add(c)
for svc in sorted(mar_svcs):
    ws_mar.cell(row=r, column=1, value=svc).font = body_font
    r += 1
r += 2

ws_mar.cell(row=r, column=1, value="Detailed Incidents").font = section_font
r += 2
mar_incs = sorted([inc for inc in grouped if inc['start'].month == 3 and inc['start'].year == 2026], key=lambda x: -x['duration'])
for idx, inc in enumerate(mar_incs, 1):
    comps = inc['components']
    dur = inc['duration']
    date_fmt = inc['start'].strftime("%B %d, %Y")
    start_fmt = inc['start'].strftime("%-I:%M %p GMT")
    end_fmt = inc['end'].strftime("%-I:%M %p GMT")
    if len(comps) > 1:
        title = f"{idx}. {date_fmt} - Multiple Outages ({dur:.0f} minutes total)"
    else:
        title = f"{idx}. {date_fmt} - Outage ({dur:.0f} minute{'s' if dur != 1 else ''})"
    ws_mar.cell(row=r, column=1, value=title).font = section_font
    r += 2
    if len(comps) > 1:
        for ci, c in enumerate(comps, 1):
            ws_mar.cell(row=r, column=1, value=f"Service {ci}: {c}").font = body_font
            r += 1
    else:
        ws_mar.cell(row=r, column=1, value=f"Service: {comps[0]}").font = body_font
        r += 1
    ws_mar.cell(row=r, column=1, value=f"Time: {start_fmt}-{end_fmt}").font = body_font
    r += 1
    ws_mar.cell(row=r, column=1, value="Impact: Brief service disruption").font = body_font
    r += 2

# ============================================================
# TAB 6: Apr 2026 (stub - in progress)
# ============================================================
ws_apr = wb.create_sheet("Apr 2026")
ws_apr.column_dimensions['A'].width = 42
ws_apr.column_dimensions['B'].width = 16
ws_apr.column_dimensions['C'].width = 18
ws_apr.column_dimensions['D'].width = 18
ws_apr.column_dimensions['E'].width = 18
ws_apr.column_dimensions['G'].width = 85

r = 1
ws_apr.cell(row=r, column=1, value="April 2026 - Monthly Detail (Partial)").font = title_font
r += 2
ws_apr.cell(row=r, column=1, value="Month in progress. Partial data through April 16, 2026. To be finalized after April 30.").font = note_font
r += 2

apr_min = 43200
apr_stats = monthly.get("2026-04", {'count':0,'minutes':0})
apr_up_partial = 1 - (apr_stats['minutes'] / apr_min) if apr_stats['minutes'] > 0 else 1

ws_apr.cell(row=r, column=1, value="Monthly Summary (Partial)").font = section_font
r += 2
for label, val in [
    ("Period:", "April 1 - 30, 2026"),
    ("Total Minutes:", f"{apr_min:,}"),
    ("Outage Count (to date):", str(apr_stats['count'])),
    ("Outage Minutes (to date):", str(int(apr_stats['minutes']))),
    ("Uptime % (partial, not final):", f"{apr_up_partial*100:.4f}%"),
    ("Status:", "IN PROGRESS, partial data through April 16, 2026"),
]:
    ws_apr.cell(row=r, column=1, value=label).font = section_font
    ws_apr.cell(row=r, column=2, value=val).font = body_font
    r += 1
r += 1

ws_apr.cell(row=r, column=1, value="Daily Outage Report (Partial)").font = section_font
r += 1
for col, h in enumerate(["Date", "Outage Count", "Outage Minutes", "Uptime Percentage"], 1):
    c = ws_apr.cell(row=r, column=col, value=h)
    c.font = hdr_font_w; c.fill = hdr_fill; c.border = thin_border; c.alignment = Alignment(horizontal="center")
r += 1

apr_daily = {d: s for d, s in daily.items() if d.month == 4 and d.year == 2026}
for d, stats in sorted(apr_daily.items(), key=lambda x: -x[1]['minutes']):
    ws_apr.cell(row=r, column=1, value=datetime(d.year, d.month, d.day)).font = body_font
    ws_apr.cell(row=r, column=1).number_format = "mm-dd-yy"
    ws_apr.cell(row=r, column=1).border = thin_border
    ws_apr.cell(row=r, column=2, value=stats['count']).font = body_font
    ws_apr.cell(row=r, column=2).border = thin_border
    ws_apr.cell(row=r, column=3, value=int(stats['minutes'])).font = body_font
    ws_apr.cell(row=r, column=3).border = thin_border
    ws_apr.cell(row=r, column=4, value=1 - (stats['minutes']/1440)).font = body_font
    ws_apr.cell(row=r, column=4).number_format = "0.00%"
    ws_apr.cell(row=r, column=4).border = thin_border
    r += 1
r += 1

ws_apr.cell(row=r, column=1, value="Services Affected (to date)").font = section_font
r += 2
apr_svcs = set()
for inc in grouped:
    if inc['start'].month == 4 and inc['start'].year == 2026:
        for c in inc['components']: apr_svcs.add(c)
for svc in sorted(apr_svcs):
    ws_apr.cell(row=r, column=1, value=svc).font = body_font
    r += 1
r += 2

ws_apr.cell(row=r, column=1, value="Detailed Incidents (to date)").font = section_font
r += 2
apr_incs = sorted([inc for inc in grouped if inc['start'].month == 4 and inc['start'].year == 2026], key=lambda x: -x['duration'])
for idx, inc in enumerate(apr_incs, 1):
    comps = inc['components']
    dur = inc['duration']
    date_fmt = inc['start'].strftime("%B %d, %Y")
    start_fmt = inc['start'].strftime("%-I:%M %p GMT")
    end_fmt = inc['end'].strftime("%-I:%M %p GMT")
    if len(comps) > 1:
        title = f"{idx}. {date_fmt} - Multiple Outages ({dur:.0f} minutes total)"
    else:
        title = f"{idx}. {date_fmt} - Outage ({dur:.0f} minute{'s' if dur != 1 else ''})"
    ws_apr.cell(row=r, column=1, value=title).font = section_font
    r += 2
    if len(comps) > 1:
        for ci, c in enumerate(comps, 1):
            ws_apr.cell(row=r, column=1, value=f"Service {ci}: {c}").font = body_font
            r += 1
    else:
        ws_apr.cell(row=r, column=1, value=f"Service: {comps[0]}").font = body_font
        r += 1
    ws_apr.cell(row=r, column=1, value=f"Time: {start_fmt}-{end_fmt}").font = body_font
    r += 1
    ws_apr.cell(row=r, column=1, value="Impact: Brief service disruption").font = body_font
    r += 2

# ============================================================
# TAB 7: 2026 Incident Data (raw)
# ============================================================
ws_raw = wb.create_sheet("2026 Incident Data")
headers_raw = ["Event Type", "Start Date Time", "End Date Time", "Service", "Duration Minutes", "Month"]
for col, h in enumerate(headers_raw, 1):
    c = ws_raw.cell(row=1, column=col, value=h)
    c.font = hdr_font_w; c.fill = hdr_fill; c.border = thin_border
all_sorted = sorted(parsed, key=lambda x: x[0], reverse=True)
for i, (sd, ed, dur, typ, comp) in enumerate(all_sorted, 2):
    ws_raw.cell(row=i, column=1, value=typ).font = body_font; ws_raw.cell(row=i,column=1).border = thin_border
    ws_raw.cell(row=i, column=2, value=sd.strftime("%Y-%m-%dT%H:%M:%SZ")).font = body_font; ws_raw.cell(row=i,column=2).border = thin_border
    ws_raw.cell(row=i, column=3, value=ed.strftime("%b %d, %Y %-I:%M %p GMT")).font = body_font; ws_raw.cell(row=i,column=3).border = thin_border
    ws_raw.cell(row=i, column=4, value=comp).font = body_font; ws_raw.cell(row=i,column=4).border = thin_border
    ws_raw.cell(row=i, column=5, value=round(dur, 1)).font = body_font; ws_raw.cell(row=i,column=5).border = thin_border
    ws_raw.cell(row=i, column=6, value=sd.strftime("%Y-%m")).font = body_font; ws_raw.cell(row=i,column=6).border = thin_border
ws_raw.column_dimensions['A'].width = 22
ws_raw.column_dimensions['B'].width = 28
ws_raw.column_dimensions['C'].width = 28
ws_raw.column_dimensions['D'].width = 55
ws_raw.column_dimensions['E'].width = 18
ws_raw.column_dimensions['F'].width = 12

# ============================================================
# TAB 8: Bucket Mapping
# ============================================================
ws_bkt = wb.create_sheet("Bucket Mapping")
map_headers = ["Component Contains", "Bucket", "Region", "Type"]
map_data = [
    ("Tenant Portal - FedRAMP", "FedRAMP Portal", "FedRAMP", "Portal"),
    ("API - FedRAMP", "FedRAMP API", "FedRAMP", "API"),
    ("Tenant Portal - AMS", "AMS Portal", "AMS", "Portal"),
    ("API - AMS", "AMS API", "AMS", "API"),
    ("Tenant Portal - EMEA", "EMEA Portal", "EMEA", "Portal"),
    ("API - EMEA", "EMEA API", "EMEA", "API"),
    ("Tenant Portal - APAC", "APAC Portal", "APAC", "Portal"),
    ("API - APAC", "APAC API", "APAC", "API"),
    ("Tenant Portal - SGP", "SGP Portal", "SGP", "Portal"),
    ("API - SGP", "SGP API", "SGP", "API"),
    ("Vulncat", "Vulncat", "Global", "Service"),
    ("eu-sast-aviator", "EU SAST Aviator", "EU", "SAST Aviator"),
    ("ams-sast-aviator", "AMS SAST Aviator", "AMS", "SAST Aviator"),
    ("Tenant Portal - EU", "EU Portal", "EU", "Portal"),
    ("API - EU", "EU API", "EU", "API"),
    ("Debricked Fortify Integration", "Debricked Fortify Integration", "Global", "Service"),
]
for col, h in enumerate(map_headers, 1):
    c = ws_bkt.cell(row=1, column=col, value=h)
    c.font = hdr_font_w; c.fill = hdr_fill; c.border = thin_border
for ri, rd in enumerate(map_data, 2):
    for ci, val in enumerate(rd, 1):
        c = ws_bkt.cell(row=ri, column=ci, value=val)
        c.font = body_font; c.border = thin_border
for i, w in enumerate([32, 22, 14, 14], 1):
    ws_bkt.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# TAB 9: Notes & Methodology
# ============================================================
ws_notes = wb.create_sheet("Notes & Methodology")
ws_notes.column_dimensions['A'].width = 120

r = 1
def write_note(row, text, bold=False):
    f = section_font if bold else body_font
    ws_notes.cell(row=row, column=1, value=text).font = f
    return row + 1

r = write_note(r, "Fortify on Demand - Uptime Report: Notes & Methodology", True)
r += 1
r = write_note(r, "ABOUT THIS FILE", True)
r = write_note(r, "This workbook is the 2026 Fortify on Demand uptime report. It is structured for monthly updates and yearly rollup.")
r = write_note(r, "It was first created in February 2026 using data scraped from https://status.fortify.com/history.")
r = write_note(r, "A prior report covering September 2024 - September 2025 was produced separately. This file starts fresh at 2026 to avoid overlap.")
r = write_note(r, "Last updated: April 16, 2026 (mid-month update; April data through Apr 16; events added: Apr 7 AMS Portal degradation, Apr 8 FedRAMP outage 3 min, Apr 9 EU SAST outage 2 min, Apr 9-10 SAST maintenance windows).")
r += 1

r = write_note(r, "DATA SOURCE", True)
r = write_note(r, "All incident data is sourced from: https://status.fortify.com/history")
r = write_note(r, "The history page lists events in reverse chronological order with: event type, date/time range (GMT), and affected service(s).")
r = write_note(r, "Each event row on the page shows one component. When a single incident affects multiple components (e.g., Portal + API),")
r = write_note(r, "both appear as separate lines with the same or near-same start time.")
r += 1

r = write_note(r, "WHAT COUNTS TOWARD UPTIME", True)
r = write_note(r, "Only 'Outage' type events are included in the uptime calculation.")
r = write_note(r, "'Maintenance' events are tracked in the raw data tab but excluded from uptime math.")
r = write_note(r, "'Service Degradation' events are excluded from uptime. The portal and services are still available during degradation,")
r = write_note(r, "so these do not represent true downtime. If a degradation event needs to be counted in the future, add it as a separate metric.")
r = write_note(r, "'Disaster Recovery Exercise' events are informational only and excluded.")
r += 1

r = write_note(r, "DE-DUPLICATION LOGIC", True)
r = write_note(r, "When the same incident hits multiple components simultaneously (e.g., FedRAMP Portal and FedRAMP API both go down at 2:45 PM),")
r = write_note(r, "these are grouped into a single incident for counting purposes. The rule:")
r = write_note(r, "  - If two or more outage events have start times within 60 seconds of each other, they are treated as one incident.")
r = write_note(r, "  - The duration used is the MAX duration across the grouped components (not the sum).")
r = write_note(r, "  - All affected component names are preserved in the detail write-up.")
r = write_note(r, "This prevents double-counting when Portal + API go down together.")
r += 1

r = write_note(r, "UPTIME FORMULA", True)
r = write_note(r, "  Monthly Uptime % = 1 - (Total Outage Minutes in Month / Total Minutes in Month)")
r = write_note(r, "  Daily Uptime %   = 1 - (Outage Minutes on Day / 1440)")
r = write_note(r, "  YTD Uptime %     = 1 - (Sum of Outage Minutes for Completed Months / Sum of Total Minutes for Completed Months)")
r = write_note(r, "")
r = write_note(r, "  Minutes per month: Jan=44640, Feb=40320, Mar=44640, Apr=43200, May=44640, Jun=43200,")
r = write_note(r, "                     Jul=44640, Aug=44640, Sep=43200, Oct=44640, Nov=43200, Dec=44640")
r = write_note(r, "  Minutes per day:   1440")
r = write_note(r, "  SLA target:        99.9% (43.8 minutes of allowed downtime per month on a 30-day month)")
r += 1

r = write_note(r, "REGIONAL BREAKOUT TAB", True)
r = write_note(r, "The 'Regional Breakout' tab provides per-component uptime percentages for every component tracked on status.fortify.com.")
r = write_note(r, "This tab does NOT use de-duplication. Each component's outage minutes are calculated independently.")
r = write_note(r, "This matches how status.fortify.com calculates and displays uptime per component.")
r = write_note(r, "Only 'Outage' events count against uptime (consistent with the executive summary methodology).")
r = write_note(r, "The tab has three sections: Core Services (Portal and API per region), SAST Aviator, and Other Services.")
r = write_note(r, "Monthly columns show Jan through Dec. The YTD column covers completed months only (same as the executive summary).")
r = write_note(r, "Added: April 2026.")
r += 1

r = write_note(r, "BUCKET MAPPING", True)
r = write_note(r, "The 'Bucket Mapping' tab defines how raw component strings from the status page map to reporting buckets.")
r = write_note(r, "Buckets group by region (FedRAMP, AMS, EMEA, APAC, SGP, EU) and type (Portal, API, SAST Aviator, Vulncat, Debricked Fortify Integration).")
r = write_note(r, "This mapping is used for per-bucket breakdowns if needed. The executive summary reports overall uptime.")
r += 1

r = write_note(r, "FILE STRUCTURE", True)
r = write_note(r, "Tab layout and what each tab contains:")
r = write_note(r, "")
r = write_note(r, "  [2026 Executive Summary]  -  The yearly roll-up. One table with all 12 months. YTD totals and SLA status.")
r = write_note(r, "                               Updated monthly. Future months show '-' until data is available.")
r = write_note(r, "                               Key observations in column G. This is the tab leadership reads.")
r = write_note(r, "")
r = write_note(r, "  [Regional Breakout]        -  Per-component uptime percentages for every status page component.")
r = write_note(r, "                               Three sections: Core Services (Portal/API by region), SAST Aviator, Other Services.")
r = write_note(r, "                               Monthly columns plus YTD. No cross-component de-duplication.")
r = write_note(r, "                               Numbers match status.fortify.com exactly.")

r = write_note(r, "  [Mon YYYY] (e.g. Jan 2026) - One tab per completed month. Contains: monthly summary block, daily outage table")
r = write_note(r, "                               (sorted worst-day-first), affected services list, and detailed incident write-ups.")
r = write_note(r, "                               Each incident includes: date, services, time range, and impact assessment.")
r = write_note(r, "                               The current (in-progress) month tab shows partial data until month-end.")
r = write_note(r, "")
r = write_note(r, "  [2026 Incident Data]       -  Raw event log. Every event (outage + maintenance) in reverse chronological order.")
r = write_note(r, "                               Columns: Event Type, Start Date Time (ISO), End Date Time, Service, Duration Minutes, Month.")
r = write_note(r, "                               This is the fact table. New rows are appended here each month.")
r = write_note(r, "")
r = write_note(r, "  [Bucket Mapping]           -  Reference table mapping component strings to region/service buckets.")
r = write_note(r, "                               Used for per-bucket breakdowns. Rarely changes unless new services are added.")
r = write_note(r, "")
r = write_note(r, "  [Notes & Methodology]      -  This tab. Calculation references, process documentation, and update instructions.")
r += 1

r = write_note(r, "HOW TO UPDATE THIS FILE EACH MONTH", True)
r = write_note(r, "Repeat these steps on or after the 1st of each new month. The goal is to finalize the prior month and set up the next.")
r = write_note(r, "")
r = write_note(r, "  1. Go to https://status.fortify.com/history and scroll to the month being finalized.")
r = write_note(r, "     Capture every Outage and Maintenance event for that month. For each event, record:")
r = write_note(r, "     start time (GMT), end time (GMT), event type, and affected component string exactly as shown on the page.")
r = write_note(r, "")
r = write_note(r, "  2. Append the new events as rows in the '2026 Incident Data' tab.")
r = write_note(r, "     Use ISO format for start time (YYYY-MM-DDTHH:MM:SSZ). Match the existing column layout.")
r = write_note(r, "")
r = write_note(r, "  3. For the month's detail tab (e.g., 'Mar 2026'):")
r = write_note(r, "     - Finalize the Monthly Summary block (update outage count, minutes, uptime %).")
r = write_note(r, "     - Build the Daily Outage Report table: aggregate outage incidents by date, sort by minutes descending.")
r = write_note(r, "     - List affected services.")
r = write_note(r, "     - Write up each incident: group simultaneous outages (start times within 60 seconds), note all affected services,")
r = write_note(r, "       time range, and a brief impact assessment.")
r = write_note(r, "     - Remember: only 'Outage' events count. Maintenance is tracked but excluded from uptime.")
r = write_note(r, "")
r = write_note(r, "  4. Create a new tab for the upcoming month (e.g., 'Apr 2026') with a placeholder/stub.")
r = write_note(r, "     It will be populated when that month closes.")
r = write_note(r, "")
r = write_note(r, "  5. Update the '2026 Executive Summary' tab:")
r = write_note(r, "     - Fill in the newly completed month's row in the Monthly Uptime Summary table (count, minutes, hours, uptime %).")
r = write_note(r, "     - Update the YTD TOTAL row to include the new month.")
r = write_note(r, "     - Refresh Key Observations in column G if anything noteworthy happened.")
r = write_note(r, "     - Update the SLA Performance block at the bottom.")
r = write_note(r, "")
r = write_note(r, "  6. The de-duplication rule (grouping events within 60 seconds) and the uptime formulas above")
r = write_note(r, "     should be applied consistently. If you are using a tool or script to generate this, point it at the")
r = write_note(r, "     status page URL, apply the same parsing logic, and follow this tab structure. The raw data tab")
r = write_note(r, "     is the single source of truth; everything else derives from it.")
r += 1

r = write_note(r, "FORMATTING CONVENTIONS", True)
r = write_note(r, "  Font: Aptos Narrow, 11pt body, 24pt title, 11pt bold for section headers")
r = write_note(r, "  Table headers: white text on dark blue (#2F5496)")
r = write_note(r, "  Uptime % format: 0.0000% (four decimals)")
r = write_note(r, "  Dates in daily tables: mm-dd-yy")
r = write_note(r, "  Month strings: YYYY-MM (e.g. 2026-01)")
r = write_note(r, "  Daily table sort: outage minutes descending (worst day first)")
r = write_note(r, "  Incident write-ups: grouped by month, sorted by duration descending within each month")
r += 1

r = write_note(r, "CONTACT / HISTORY", True)
r = write_note(r, "First created: February 19, 2026")
r = write_note(r, "Last updated: April 16, 2026")
r = write_note(r, "Created by: Chance Bonner (cbonner@opentext.com)")
r = write_note(r, "Source reference: Previous 2025 report was a separate deliverable; this file does not overlap with it.")
r = write_note(r, "Automation potential: The monthly update process follows a repeatable pattern and could be scripted.")
r = write_note(r, "The status page (status.fortify.com/history) is the canonical data source for all incident records.")

# Revised date note on 2026 Incident Data sheet
revised_date = datetime.now().strftime("%B %d, %Y")
last_data_row = len(all_sorted) + 2  # header row + data rows + 1
ws_raw.cell(row=last_data_row, column=1, value=f"Report revised: {revised_date}").font = note_font

# Save
output = os.path.join(os.path.dirname(os.path.abspath(__file__)), "FoD_Uptime_Report_2026.xlsx")
wb.save(output)
print(f"Saved: {output}")
print(f"Sheets: {wb.sheetnames}")
print(f"Jan: {jan_stats['count']} outages, {jan_stats['minutes']:.0f} min, {jan_up*100:.4f}%")
print(f"Feb: {feb_stats['count']} outages, {feb_stats['minutes']:.0f} min, {feb_up*100:.4f}%")
print(f"Mar: {mar_stats['count']} outages, {mar_stats['minutes']:.0f} min, {mar_up*100:.4f}%")
print(f"YTD: {ytd_outage_count} outages, {ytd_outage_min:.0f} min, {ytd_uptime*100:.4f}%")
print(f"Raw rows: {len(parsed)}")
print(f"De-duped outage incidents: {len(grouped)}")
